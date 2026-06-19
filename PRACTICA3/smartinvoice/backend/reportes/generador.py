import csv
import os
from datetime import datetime
from typing import List

REPORTS_DIR = os.getenv("REPORTS_DIR", "/app/reports")


def _asegurar_directorio():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _nombre_archivo(formato: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"reporte_{timestamp}.{formato}"


def generar_csv(facturas: List) -> str:
    """Genera un reporte CSV con las facturas proporcionadas. Retorna la ruta del archivo."""
    _asegurar_directorio()
    nombre = _nombre_archivo("csv")
    ruta = os.path.join(REPORTS_DIR, nombre)

    with open(ruta, "w", newline="", encoding="utf-8") as f:
        escritor = csv.writer(f)
        escritor.writerow([
            "ID", "Numero Factura", "Fecha", "Proveedor", "NIT",
            "Subtotal", "IVA", "Total", "Estado", "Fecha Carga"
        ])
        for fac in facturas:
            escritor.writerow([
                fac.id,
                fac.numero_factura or "",
                fac.fecha_factura or "",
                fac.proveedor_nombre or "",
                fac.proveedor_nit or "",
                fac.subtotal,
                fac.impuesto,
                fac.total,
                fac.estado,
                fac.fecha_carga.strftime("%d/%m/%Y %H:%M") if fac.fecha_carga else "",
            ])

    return ruta


def generar_excel(facturas: List) -> str:
    """Genera un reporte Excel con las facturas proporcionadas. Retorna la ruta del archivo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    _asegurar_directorio()
    nombre = _nombre_archivo("xlsx")
    ruta = os.path.join(REPORTS_DIR, nombre)

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    encabezados = [
        "ID", "Numero Factura", "Fecha", "Proveedor", "NIT",
        "Subtotal", "IVA", "Total", "Estado", "Fecha Carga"
    ]

    estilo_encabezado = Font(bold=True, color="FFFFFF")
    fondo_encabezado = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    for col_idx, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_idx, value=encabezado)
        celda.font = estilo_encabezado
        celda.fill = fondo_encabezado
        celda.alignment = Alignment(horizontal="center")

    for fila_idx, fac in enumerate(facturas, 2):
        ws.cell(row=fila_idx, column=1, value=fac.id)
        ws.cell(row=fila_idx, column=2, value=fac.numero_factura or "")
        ws.cell(row=fila_idx, column=3, value=fac.fecha_factura or "")
        ws.cell(row=fila_idx, column=4, value=fac.proveedor_nombre or "")
        ws.cell(row=fila_idx, column=5, value=fac.proveedor_nit or "")
        ws.cell(row=fila_idx, column=6, value=fac.subtotal)
        ws.cell(row=fila_idx, column=7, value=fac.impuesto)
        ws.cell(row=fila_idx, column=8, value=fac.total)
        ws.cell(row=fila_idx, column=9, value=fac.estado)
        ws.cell(row=fila_idx, column=10, value=fac.fecha_carga.strftime("%d/%m/%Y %H:%M") if fac.fecha_carga else "")

    for col in ws.columns:
        max_len = max(len(str(celda.value or "")) for celda in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    resumen_fila = len(facturas) + 3
    ws.cell(row=resumen_fila, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=resumen_fila, column=6, value=sum(f.subtotal for f in facturas)).font = Font(bold=True)
    ws.cell(row=resumen_fila, column=7, value=sum(f.impuesto for f in facturas)).font = Font(bold=True)
    ws.cell(row=resumen_fila, column=8, value=sum(f.total for f in facturas)).font = Font(bold=True)

    wb.save(ruta)
    return ruta


def generar_pdf(facturas: List) -> str:
    """Genera un reporte PDF con tabla de facturas. Retorna la ruta del archivo."""
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch

    _asegurar_directorio()
    nombre = _nombre_archivo("pdf")
    ruta = os.path.join(REPORTS_DIR, nombre)

    doc = SimpleDocTemplate(ruta, pagesize=landscape(letter))
    estilos = getSampleStyleSheet()
    elementos = []

    titulo = Paragraph("Reporte de Facturas Procesadas - SmartInvoice", estilos["Title"])
    elementos.append(titulo)
    elementos.append(Spacer(1, 0.2 * inch))

    fecha_gen = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", estilos["Normal"])
    elementos.append(fecha_gen)
    elementos.append(Spacer(1, 0.2 * inch))

    total_general = Paragraph(f"Total de facturas: {len(facturas)}", estilos["Normal"])
    elementos.append(total_general)
    elementos.append(Spacer(1, 0.3 * inch))

    datos = [["ID", "Num. Factura", "Fecha", "Proveedor", "NIT", "Subtotal", "IVA", "Total", "Estado"]]
    for fac in facturas:
        datos.append([
            str(fac.id),
            fac.numero_factura or "-",
            fac.fecha_factura or "-",
            (fac.proveedor_nombre or "-")[:25],
            fac.proveedor_nit or "-",
            f"Q{fac.subtotal:.2f}",
            f"Q{fac.impuesto:.2f}",
            f"Q{fac.total:.2f}",
            fac.estado,
        ])

    suma_total = sum(f.total for f in facturas)
    datos.append(["", "", "", "", "TOTAL GENERAL", "", "", f"Q{suma_total:.2f}", ""])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F3F4")]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D5DBDB")),
    ]))

    elementos.append(tabla)
    doc.build(elementos)
    return ruta
